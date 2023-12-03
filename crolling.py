from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pytube import YouTube
import datetime
import concurrent.futures

xlsk_dir = './YouTube.xlsx'

# 검색어를 입력하고 Selenium을 사용해서 웹 브라우저를 열어줍니다.
search_word = input("YouTube 검색어를 입력하세요 : ")
search_word = search_word.replace(' ', '+')

url = f"https://www.youtube.com/results?search_query={search_word}"
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)
driver.get(url)
time.sleep(3)

# 유튜브 페이지는 스크롤을 해서 다음 페이지를 읽어들입니다.
# 페이지 끝까지 스크롤하는 코드입니다.

SCROLL_PAUSE_TIME = 1.5
last_height = driver.execute_script("return document.documentElement.scrollHeight")

while True:

    driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")

    time.sleep(SCROLL_PAUSE_TIME)

    new_height = driver.execute_script("return document.documentElement.scrollHeight;")

    if new_height == last_height:
        break

    last_height = new_height


# 해당 경로에서 파일을 불러옵니다.
# 파일이 없다면 파일을 하나 새로 만들어 줍니다.

try:

    wb = load_workbook(xlsk_dir)

    check_new_file = 0

except:

    wb = Workbook()

    check_new_file = 1

# 만약 파일을 새로 만들었다면, 새로 만든 파일에서 시트가 하나 남을때까지 지워줍니다.

# 모든 시트를 지우고, 첫번째 시트 이름을 [ 검색어 + 날짜 + 시간 ] 형태로 시트명을 입력합니다.

# 엑셀의 시트명으로 [ : ]을 사용할 수 없기에 시간역시 [ - ]로 구분했습니다.

# 만약 기존에 있던 파일을 불러왔다면, 동일한 방법으로 시트를 추가해줍니다.

if check_new_file == 1:

    while len(wb.sheetnames) > 1:
        wb.remove(wb.sheetnames[len(wb.sheetnames) - 1])
    ws = wb.active
    ws.title = search_word.replace('[', '').replace(']', '') + time.strftime("(%Y-%m-%d %H-%M-%S)")

else:
    ws = wb.create_sheet(search_word.replace('[', '').replace(']', '') + time.strftime("(%Y-%m-%d %H-%M-%S)"))

    # 엑셀파일에 [ 제목, URL, 조회수, 업데이트 날짜, 길이 ]라는 제목을 만들어줍니다.
    # append() 함수를 사용해서 값을 추가하면 위에서 부터 오른쪽으로 차례로 값을 입력합니다.
    # 다시 append() 함수를 사용하면 다음 줄의 처음부터 오른쪽으로 차례로 값을 입력합니다.
ws.append(['제목', 'URL', "조회수", "업데이트 날짜", "길이"])

    # Selenium을 사용해 유튜브의 제목과 URL을 크롤링합니다.
    # append() 함수를 사용해 제목과 URL을 엑셀 파일에 입력합니다.

titles = driver.find_elements(By.CSS_SELECTOR, "#dismissible.style-scope.ytd-video-renderer")

for title in titles:
    main_title = title.find_element(By.CSS_SELECTOR, "#video-title").get_property("title")
    search_word = search_word.replace('+', ' ')

    # 검색어로 시작하지 않는 제목은 건너뜁니다.
    if not main_title.startswith(search_word):
        continue

    tube_url = title.find_element(By.CSS_SELECTOR, "#video-title").get_property("href")

    ws.append([main_title, tube_url])


# row의 총 갯수를 불러와서 저장합니다.

maxrow = ws.max_row

# count_row 변수와 k 변수는 진행상황을 print 하기위해 선언한 변수입니다.
# 아래에서 확인이 가능합니다.

count_row = maxrow - 1
k = 1

# 여기부터는 엑셀파일의 URL을 차례로 불러온 후
# pytube 모듈을 사용하여 해당 URL의 유튜브 영상의 조회수, 업데이트 날짜, 길이를 불러옵니다.

for i in range(2, maxrow + 1):
    url = ws.cell(row=i, column=2).value
    tube = YouTube(url)
    view = tube.views  # 조회수 불러오기
    update_dates = str(tube.publish_date)  # 업데이트 날짜 불러와서 str 형태로 변환
    update_date = update_dates.split(" ")  # 업데이트 날짜 string을 공백으로 분리
    # 업데이트 날짜는 0000-00-00 00-00-00 형태인데 앞의 날짜부분만 추출하려 합니다.
    length_second = int(tube.length)  # 유튜브 영상의 길이를 불러옴(문자 형태인것 같아서 int로 변환) / 문자가 아닐 수도...
    length = str(datetime.timedelta(seconds=length_second))  # datetime 모듈의 timedelta 함수로 초를 시:분:초 로 변환

    ws.cell(row=i, column=3).value = view
    ws.cell(row=i, column=3).number_format = "#,##0"  # 조회수를 세자리마다 [ , ]를 찍어 구분
    ws.cell(row=i, column=4).value = update_date[0]  # 추출한 날짜 중 앞의 날짜부분만 추출
    ws.cell(row=i, column=4).alignment = Alignment(horizontal="center", vertical="center")  # 가운데 정렬
    ws.cell(row=i, column=5).value = length
    ws.cell(row=i, column=5).alignment = Alignment(horizontal="center", vertical="center")

    print(f"총 {count_row}개 중 {k}번째 완료")  # 진행상황을 명령창에 표시

    k = k + 1


# 파일 저장하고 닫기

wb.save(xlsk_dir)
wb.close()